import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import date, datetime
from tkcalendar import DateEntry
import os

# --- Load Customers from Excel File ---
def load_customers_from_excel(file_path="customer_data.xlsx"):
    """Loads customer data from an Excel file."""
    customers = {}
    try:
        if not os.path.exists(file_path):
            messagebox.showwarning("customer_data.xlsx Not Found, Creating a new file named customer_data.xlsx", 
                f"Customer database file '{file_path}' not found.\nUsing empty customer list.")
            return customers
            
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Read data starting from row 2 (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Check if Customer Key exists
                customer_key = row[0]
                customers[customer_key] = {
                    "name": row[1] if row[1] else "",
                    "gst": row[2] if row[2] else "",
                    "address": row[3] if row[3] else "",
                    "phone": row[4] if len(row) > 4 and row[4] else "",
                    "email": row[5] if len(row) > 5 and row[5] else "",
                }
        wb.close()
    except Exception as e:
        messagebox.showerror("Error Loading Customers", 
            f"Failed to load customer data from '{file_path}'.\nError: {e}")
    
    return customers

def save_customer_to_excel(customer_key, customer_data, file_path="customer_data.xlsx"):
    """Saves a new customer to the Excel file."""
    try:
        if not os.path.exists(file_path):
            # Create new file with headers if it doesn't exist
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Customers"

            # Create headers
            headers = ["Customer Key", "Customer Name", "GST Number", "Address", "Phone", "Email"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 45
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 28
        else:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

        # Find the next empty row
        next_row = ws.max_row + 1

        # Add the new customer data
        ws.cell(row=next_row, column=1, value=customer_key)
        ws.cell(row=next_row, column=2, value=customer_data.get("name", ""))
        ws.cell(row=next_row, column=3, value=customer_data.get("gst", ""))
        ws.cell(row=next_row, column=4, value=customer_data.get("address", ""))
        ws.cell(row=next_row, column=5, value=customer_data.get("phone", ""))
        ws.cell(row=next_row, column=6, value=customer_data.get("email", ""))

        # Save the workbook
        wb.save(file_path)
        wb.close()
        return True
    except Exception as e:
        messagebox.showerror("Error Saving Customer", 
            f"Failed to save customer data to '{file_path}'.\nError: {e}")
        return False

# Load customers at startup
SAVED_PARTIES = load_customers_from_excel()

class InvoiceGeneratorApp:
    def __init__(self, master):
        self.master = master
        master.title("🧾 Tax Invoice Generator")
        master.geometry("800x700")
        # Maximize the window
        master.state('zoomed')
        
        # --- Data storage for Item details ---
        self.items_data = []
        
        # --- Setup the main notebook (tabs) ---
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(pady=10, padx=10, expand=True, fill="both")
        
        self.page1 = ttk.Frame(self.notebook)
        
        self.notebook.add(self.page1, text="Buyer Details & Items")
        
        self._setup_buyer_details_frame(self.page1)
        self._setup_items_frame(self.page1)

    ## ----------------- PAGE 1: BUYER DETAILS -----------------

    def _setup_buyer_details_frame(self, parent):
        """Creates the frame for Buyer and General Invoice Details."""
        frame = ttk.LabelFrame(parent, text="Buyer & Invoice Details")
        frame.pack(fill="x", padx=10, pady=5)
        
        # Grid Configuration for the frame
        for i in range(4):
            frame.columnconfigure(i, weight=1)

        # 1. Saved Party Dropdown
        ttk.Label(frame, text="Load Saved Party:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        party_keys = ["(New Party)"] + list(SAVED_PARTIES.keys())
        self.party_var = tk.StringVar(frame)
        self.party_var.set(party_keys[0]) # Default value
        
        self.party_dropdown = ttk.Combobox(frame, textvariable=self.party_var, values=party_keys, state="readonly")
        self.party_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.party_dropdown.bind("<<ComboboxSelected>>", self._load_saved_party)

        # 2. Buyer Details Entry Fields
        self.buyer_entries = {}
        fields = [
            ("Buyer's Name", "name", 1), 
            ("GST No.", "gst", 2), 
            ("Party's Address", "address", 3),
            ("Customer Phone", "phone", 4),
            ("Customer Email", "email", 5)
        ]

        for i, (label_text, key, row_num) in enumerate(fields):
            ttk.Label(frame, text=f"{label_text}:").grid(row=row_num, column=0, padx=5, pady=2, sticky="w")
            entry = ttk.Entry(frame)
            entry.grid(row=row_num, column=1, padx=5, pady=2, sticky="ew")
            self.buyer_entries[key] = entry

        # 3. Date Fields with Calendar
        self.date_entries = {}
        date_fields = [
            ("Sale Date", "sale_date", 1), 
            ("Delivery Date", "delivery_date", 2)
        ]
        
        # Default dates
        default_sale_date = date.today()
        
        for i, (label_text, key, row_num) in enumerate(date_fields):
            ttk.Label(frame, text=f"{label_text}:").grid(row=row_num, column=2, padx=5, pady=2, sticky="w")
            date_entry = DateEntry(frame, width=18, background='darkblue',
                                  foreground='white', borderwidth=2, 
                                  date_pattern='dd-mm-yyyy')
            date_entry.grid(row=row_num, column=3, padx=5, pady=2, sticky="ew")
            self.date_entries[key] = date_entry
            if key == "sale_date":
                date_entry.set_date(default_sale_date)

    def _load_saved_party(self, event):
        """Loads details from SAVED_PARTIES into entry fields."""
        selected_key = self.party_var.get()
        if selected_key in SAVED_PARTIES:
            details = SAVED_PARTIES[selected_key]
            for key, entry in self.buyer_entries.items():
                entry.delete(0, tk.END)
                entry.insert(0, details.get(key, ""))
        elif selected_key == "(New Party)":
             for entry in self.buyer_entries.values():
                entry.delete(0, tk.END)

    ## ----------------- PAGE 1: ITEMS ENTRY -----------------

    def _setup_items_frame(self, parent):
        """Creates the frame for adding items and the Treeview for display."""
        frame = ttk.LabelFrame(parent, text="Goods Details")
        frame.pack(fill="both", padx=10, pady=5, expand=True)
        
        # 1. Item Entry fields
        entry_frame = ttk.Frame(frame)
        entry_frame.pack(fill="x", padx=5, pady=5)

        # Make columns responsive
        for i in range(10):
            entry_frame.columnconfigure(i, weight=1)
        self.item_entries = {}

        # HSN/SAC Code (Optional)
        ttk.Label(entry_frame, text="HSN/SAC Code").grid(row=0, column=0, padx=5, sticky="ew")
        hsn_entry = ttk.Entry(entry_frame)
        hsn_entry.grid(row=0, column=1, padx=5, sticky="ew")
        self.item_entries["hsn"] = hsn_entry

        # Description
        ttk.Label(entry_frame, text="Description").grid(row=0, column=2, padx=5, sticky="ew")
        desc_entry = ttk.Entry(entry_frame)
        desc_entry.grid(row=0, column=3, padx=5, sticky="ew")
        desc_entry.bind("<Return>", lambda event: self._add_item())
        self.item_entries["description"] = desc_entry

        # Quantity
        ttk.Label(entry_frame, text="Quantity").grid(row=0, column=4, padx=5, sticky="ew")
        qty_entry = ttk.Entry(entry_frame)
        qty_entry.grid(row=0, column=5, padx=5, sticky="ew")
        qty_entry.bind("<Return>", lambda event: self._add_item())
        self.item_entries["quantity"] = qty_entry

        # Rate
        ttk.Label(entry_frame, text="Rate").grid(row=0, column=6, padx=5, sticky="ew")
        rate_entry = ttk.Entry(entry_frame)
        rate_entry.grid(row=0, column=7, padx=5, sticky="ew")
        rate_entry.bind("<Return>", lambda event: self._add_item())
        self.item_entries["rate"] = rate_entry

        # CGST
        ttk.Label(entry_frame, text="CGST (%)").grid(row=0, column=8, padx=5, sticky="ew")
        gst_frame = ttk.Frame(entry_frame)
        gst_frame.grid(row=0, column=9, padx=5, sticky="ew")
        gst_entry = ttk.Entry(gst_frame)
        gst_entry.pack(side="left", fill="x", expand=True)
        gst_entry.bind("<Return>", lambda event: self._add_item())
        self.item_entries["gst"] = gst_entry
        ttk.Label(gst_frame).pack(side="left")

        # SGST
        ttk.Label(entry_frame, text="SGST (%)").grid(row=1, column=8, padx=5, sticky="ew")
        sgst_frame = ttk.Frame(entry_frame)
        sgst_frame.grid(row=1, column=9, padx=5, sticky="ew")
        sgst_entry = ttk.Entry(sgst_frame)
        sgst_entry.pack(side="left", fill="x", expand=True)
        sgst_entry.bind("<Return>", lambda event: self._add_item())
        self.item_entries["sgst"] = sgst_entry
        ttk.Label(sgst_frame).pack(side="left")

        # 2. Treeview for displaying added items
        self.tree = ttk.Treeview(frame, columns=("HSN/SAC","Description", "Qty", "Rate", "CGST", "SGST", "Total"), show="headings", height=10)
        self.tree.heading("HSN/SAC", text="HSN/SAC Code")
        self.tree.column("HSN/SAC", width=100, anchor="center")
        self.tree.heading("Description", text="Description")
        self.tree.column("Description", width=200, anchor="w")
        self.tree.heading("Qty", text="Quantity")
        self.tree.column("Qty", width=80, anchor="center")
        self.tree.heading("Rate", text="Rate")
        self.tree.column("Rate", width=100, anchor="e")
        self.tree.heading("CGST", text="CGST")
        self.tree.column("CGST", width=80, anchor="center")
        self.tree.heading("SGST", text="SGST")
        self.tree.column("SGST", width=80, anchor="center")
        self.tree.heading("Total", text="Total Amount")
        self.tree.column("Total", width=120, anchor="e")
        self.tree.pack(fill="both", padx=5, pady=5, expand=True)

        # Buttons frame at the bottom
        buttons_frame = ttk.Frame(frame)
        buttons_frame.pack(pady=5, anchor="e")
        
        ttk.Button(buttons_frame, text="Remove Selected Item", command=self._remove_item).pack(side="left", padx=5)
        ttk.Button(buttons_frame, text="Generate Invoice (Excel)", command=self._generate_invoice).pack(side="left", padx=5)
        
    def _add_item(self):
        """Validates input and adds an item to the items_data list and Treeview."""
        try:
            hsn = self.item_entries["hsn"].get().strip()
            description = self.item_entries["description"].get().strip()
            quantity = float(self.item_entries["quantity"].get())
            rate = float(self.item_entries["rate"].get())
            
            # Parse CGST percentage
            cgst_input = self.item_entries["gst"].get().strip()
            cgst_percent = float(cgst_input.rstrip('%')) if cgst_input else 0.0
            
            # Parse SGST percentage
            sgst_input = self.item_entries["sgst"].get().strip()
            sgst_percent = float(sgst_input.rstrip('%')) if sgst_input else 0.0

            if not description or quantity <= 0 or rate <= 0 or cgst_percent < 0 or sgst_percent < 0:
                messagebox.showerror("Input Error", "All fields must be filled, and Quantity/Rate/CGST/SGST must be valid positive numbers.")
                return

            # Calculate amounts: CGST and SGST are percentages of the item total (Quantity × Rate)
            item_subtotal = quantity * rate
            cgst_amount = (item_subtotal * cgst_percent) / 100.0
            sgst_amount = (item_subtotal * sgst_percent) / 100.0
            total_tax = cgst_amount + sgst_amount
            total_with_tax = item_subtotal + total_tax

            item = {
                "hsn": hsn,
                "description": description,
                "quantity": quantity,
                "rate": rate,
                "cgst_percent": cgst_percent,
                "sgst_percent": sgst_percent,
                "cgst_amount": cgst_amount,
                "sgst_amount": sgst_amount,
                "total": item_subtotal,
            }
            self.items_data.append(item)

            # Insert into Treeview
            self.tree.insert("", "end", values=(
                hsn,
                description,
                f"{quantity:.2f}",
                f"{rate:.2f}",
                f"{cgst_amount:.2f}",
                f"{sgst_amount:.2f}",
                f"{total_with_tax:.2f}"
            ))

            # Clear fields after adding
            for key in self.item_entries:
                self.item_entries[key].delete(0, tk.END)
            self.item_entries["description"].focus()

        except ValueError:
            messagebox.showerror("Input Error", "Quantity, Rate, CGST, and SGST must be valid numbers. Enter percentages like '18' or '18%'.")

    def _remove_item(self):
        """Removes the selected item from the Treeview and the items_data list."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Selection Error", "Please select an item to remove.")
            return

        # Get the index of the selected item in the Treeview
        item_id = selected_item[0]
        item_index = self.tree.index(item_id)
        
        # Remove from the Python list
        if 0 <= item_index < len(self.items_data):
            self.items_data.pop(item_index)
            
        # Remove from the Treeview
        self.tree.delete(item_id)

    ## ----------------- INVOICE GENERATION -----------------
    def _get_all_input_data(self):
        """Collects all data from GUI inputs for validation and generation."""
        
        # 1. Buyer Data
        party_details = {
            key: entry.get().strip() for key, entry in self.buyer_entries.items()
        }
        
        # 2. General Invoice Data - Format dates as DD-MM-YYYY
        invoice_details = {
            key: entry.get_date().strftime('%d-%m-%Y') for key, entry in self.date_entries.items()
        }
        
        # 3. Items Data
        items = self.items_data
        
        # --- Basic Validation ---
        if not all(party_details.values()):
            return None, "All Buyer details must be filled out."
        if not all(invoice_details.values()):
            return None, "Both Sale Date and Delivery Date must be filled out."
        if not items:
            return None, "At least one item must be added to the invoice."

        return {
            "party": party_details,
            "invoice": invoice_details,
            "items": items
        }, None

    def _generate_invoice(self):
        """Gathers data, validates, and calls the Excel generation function."""
        data, error = self._get_all_input_data()
        
        if error:
            messagebox.showerror("Generation Error", error)
            return
            
        try:
            # Check if this is a new party and save it
            party_name = data["party"]["name"]
            is_new_party = self.party_var.get() == "(New Party)"
            
            if is_new_party and party_name:
                # Generate customer key from name
                customer_key = party_name.split()[0].upper().replace(".", "").replace(",", "")
                
                # Check if customer already exists
                global SAVED_PARTIES
                if customer_key not in SAVED_PARTIES:
                    # Save to Excel
                    if save_customer_to_excel(customer_key, data["party"]):
                        # Reload customers
                        SAVED_PARTIES = load_customers_from_excel()
                        # Update dropdown values
                        party_keys = ["(New Party)"] + list(SAVED_PARTIES.keys())
                        self.party_dropdown['values'] = party_keys
                        messagebox.showinfo("Customer Saved", 
                            f"New customer '{party_name}' has been saved to the database.")
            
            # Call the Excel generation function
            filename = self._generate_invoice_excel(data["party"], data["invoice"], data["items"])
            messagebox.showinfo("Success!", f"Invoice successfully generated as:\n**{filename}**\n\nFile saved in the current directory.")
            
        except Exception as e:
            messagebox.showerror("Critical Error", f"An error occurred during file generation: {e}")

    ## ----------------- EXCEL GENERATION LOGIC -----------------
    
    def _generate_invoice_excel(self, party_details, invoice_details, items):
        """Generates the invoice in an Excel file (Adapted from previous script)."""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tax Invoice"

        # --- Styles (Basic) ---
        heading_font = Font(name='Calibri', size=16, bold=True)
        header_font = Font(name='Calibri', size=11, bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
        fill_header = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # --- Invoice Header (Your Company Details) ---
        ws['A1'] = "Anant Enterprises"
        ws['A1'].font = heading_font
        ws['A2'] = "18/560, New industrial estate"
        ws['A3'] = "Ichalkaranji Opp. ASC College, Kolhapur, Maharashtra, 416115"
        ws['A4'] = "GSTIN: 27FQLPP6106G1ZK"

        # --- Merge cells for logo ---
        ws.merge_cells('E1:G1')
        ws.merge_cells('E2:G2')
        ws.merge_cells('E3:G3')

        # --- Insert logo image ---
        logo_path = os.path.join(os.getcwd(), "logo.jpg")
        if os.path.exists(logo_path):
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            # Resize the image to fit the merged cells (approximate)
            img.width = 250  # pixels
            img.height = 80  # pixels
            ws.add_image(img, 'E1')
        
        # --- Buyer Details ---
        ws['A6'] = "BILL TO:"
        ws['A6'].font = header_font
        ws['A7'] = f"Party Name: {party_details['name']}"
        ws['A8'] = f"GST No.: {party_details['gst']}"
        ws['A9'] = f"Address: {party_details['address']}"
        ws['A10'] = f"Phone: {party_details.get('phone', '')}"
        ws['A11'] = f"Email: {party_details.get('email', '')}"
        
        # --- Date Details ---
        ws['G6'] = "INVOICE DATE:"
        ws['G6'].font = header_font
        ws['H6'] = invoice_details['sale_date']
        ws['G7'] = "DELIVERY DATE:"
        ws['G7'].font = header_font
        ws['H7'] = invoice_details['delivery_date']
        
        # --- Table Headers (Row 12) ---
        headers = ["Sr.", "HSN/SAC Code", "Description of Goods", "Quantity", "Rate", "CGST", "SGST", "Total (Incl. Tax)"]
        col_start = 1
        for i, header in enumerate(headers):
            col = col_start + i
            cell = ws.cell(row=13, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = fill_header
            cell.border = border
        
        # --- Table Data ---
        row = 14
        total_invoice_amount = 0

        for i, item in enumerate(items):
            ws[f'A{row}'] = i + 1
            cell = ws[f'A{row}']
            cell.alignment = Alignment(horizontal='center', vertical='center') # Center align serial number
            ws[f'B{row}'] = item['hsn']
            # Apply center alignment and wrap text to HSN/SAC Code
            hsn_cell = ws[f'B{row}']
            hsn_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[f'C{row}'] = item['description']
            ws[f'D{row}'] = item['quantity']
            cell = ws[f'C{row}']
            cell.alignment = Alignment(horizontal='center', vertical='center') # Center align quantity
            ws[f'E{row}'] = item['rate']
            # Display CGST and SGST as percentages with % symbol
            ws[f'F{row}'] = f"{item['cgst_percent']:.2f}%"
            ws[f'G{row}'] = f"{item['sgst_percent']:.2f}%"

            # Set wrap text, center and middle alignment for description cell
            desc_cell = ws[f'C{row}']
            desc_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Calculate total with CGST and SGST added
            total_with_tax = item['total'] + item['cgst_amount'] + item['sgst_amount']
            ws[f'H{row}'] = total_with_tax

            # Formatting and border
            for col_idx in range(1, 9):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = border
                if col_idx >= 4:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            total_invoice_amount += total_with_tax
            row += 1

        # --- Calculate Total GST ---
        total_gst = sum(item['cgst_amount'] + item['sgst_amount'] for item in items)

        # --- Summary & Totals ---
        # Total GST row
        ws[f'G{row + 1}'] = "Total GST:"
        ws[f'G{row + 1}'].font = header_font
        ws[f'H{row + 1}'] = total_gst
        ws[f'H{row + 1}'].font = header_font
        ws[f'H{row + 1}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Total (Incl. Tax) row
        ws[f'G{row + 2}'] = "TOTAL (Incl. Tax):"
        ws[f'G{row + 2}'].font = header_font
        ws[f'H{row + 2}'] = total_invoice_amount
        ws[f'H{row + 2}'].font = header_font
        ws[f'H{row + 2}'].alignment = Alignment(horizontal='center', vertical='center')

        # --- Signature Section ---
        signature_start_row = row + 1
        ws[f'D{signature_start_row}'] = "For Anant Enterprises"
        # ws[f'E{signature_start_row}'].alignment = Alignment(horizontal='center', vertical='center') 
        ws[f'D{signature_start_row}'].font = header_font
        
        # Add Authority Signatory after some space
        signatory_row = signature_start_row + 4
        ws[f'D{signatory_row}'] = "Authority Signatory"
        # desc_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{signatory_row}'].font = header_font

        #Add Bank Details
        bank_details_row = row + 1
        ws[f'A{bank_details_row}'] = "Bank Details:"
        ws[f'A{bank_details_row}'].font = header_font
        ws[f'A{bank_details_row + 1}'] = "A/C: ANANT ENTERPRISES"
        ws[f'A{bank_details_row + 2}'] = "A/C No.: 50200104022360"
        ws[f'A{bank_details_row + 3}'] = "IFSC Code: HDFC0007957"
        ws[f'A{bank_details_row + 4}'] = "Branch: Ichalkaranji"
        
        # --- Column Widths for readability ---
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 13

        # --- Page Layout: Scale to Fit (1 page width) ---
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.print_options.horizontalCentered = True
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1

        # --- Save the file ---
        # Ensure the filename is safe and unique
        party_name_safe = party_details['name'].split()[0].replace('.', '').upper()
        filename = os.path.join(os.getcwd(), f"Invoice_{party_name_safe}_{date.today()}.xlsx")
        wb.save(filename)
        return filename

# --- Main Tkinter Loop ---
if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceGeneratorApp(root)
    root.mainloop()
